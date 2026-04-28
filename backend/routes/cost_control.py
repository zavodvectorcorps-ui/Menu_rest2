"""
Контроль цен и маржинальности
- Загрузка себестоимости из XLSX/CSV файлов
- Импорт из Caffesta POS (cost_price в тех.картах)
- Расчёт маржинальности: пороги (item > category > default)
- Telegram-алерты при падении маржи
"""
import csv
import io
import re
from datetime import datetime, timezone, timedelta
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel

from auth import check_restaurant_access, get_current_user, ensure_can_write_system
from database import db
from helpers import get_or_create_settings
from services.caffesta import get_caffesta_config, caffesta_get_balances

router = APIRouter()


# ============ Helpers ============
def _normalize_name(s: str) -> str:
    if not s:
        return ""
    s = s.lower().strip()
    s = re.sub(r"[^\w\s]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def _get_effective_threshold(item, category, settings) -> int:
    """Cascade: item.margin_threshold > category.margin_threshold > settings.margin_threshold_default"""
    if item and item.get("margin_threshold") is not None:
        return int(item["margin_threshold"])
    if category and category.get("margin_threshold") is not None:
        return int(category["margin_threshold"])
    return int((settings or {}).get("margin_threshold_default", 30))


def _compute_margin(price, cost):
    """Return (margin_pct, status). Status: 'ok' | 'warning' | 'critical' | 'no-cost'"""
    if cost is None or cost <= 0:
        return None, "no-cost"
    if price is None or price <= 0:
        return None, "no-price"
    margin = (price - cost) / price * 100
    return round(margin, 1), None


# ============ Parse upload ============
def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text), delimiter=",")
    rows = list(reader)
    if not rows:
        return []
    # Try to detect delimiter if only 1 column
    if len(rows[0]) == 1 and ";" in rows[0][0]:
        reader = csv.reader(io.StringIO(text), delimiter=";")
        rows = list(reader)
    if len(rows) < 2:
        return []
    return _rows_to_items(rows)


def _parse_xlsx(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    return _rows_to_items(rows)


def _rows_to_items(rows: list) -> list[dict]:
    """First row = headers. Find columns matching 'name' and 'cost'."""
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip().lower() for h in rows[0]]
    name_idx = None
    cost_idx = None
    for i, h in enumerate(headers):
        if name_idx is None and any(k in h for k in ["назв", "name", "блюдо", "товар", "product"]):
            name_idx = i
        if cost_idx is None and any(k in h for k in ["себест", "cost", "price", "закуп", "цена"]):
            cost_idx = i
    if name_idx is None:
        name_idx = 0
    if cost_idx is None:
        cost_idx = 1
    result = []
    for r in rows[1:]:
        if name_idx >= len(r) or cost_idx >= len(r):
            continue
        name = str(r[name_idx] or "").strip()
        cost = _parse_float(r[cost_idx])
        if name and cost is not None:
            result.append({"name": name, "cost": cost})
    return result


# ============ Matching ============
async def _match_and_update(restaurant_id: str, entries: list[dict], source: str) -> dict:
    """Match by Caffesta ID first (if provided), then by normalized name."""
    items = await db.menu_items.find({"restaurant_id": restaurant_id}, {"_id": 0}).to_list(10000)
    by_name = {_normalize_name(i["name"]): i for i in items}
    by_cfid = {i.get("caffesta_product_id"): i for i in items if i.get("caffesta_product_id")}

    matched = 0
    unmatched = []
    now = datetime.now(timezone.utc).isoformat()

    for entry in entries:
        target = None
        if entry.get("caffesta_product_id") and entry["caffesta_product_id"] in by_cfid:
            target = by_cfid[entry["caffesta_product_id"]]
        elif entry.get("name"):
            target = by_name.get(_normalize_name(entry["name"]))

        if target:
            await db.menu_items.update_one(
                {"id": target["id"]},
                {"$set": {
                    "cost_price": entry["cost"],
                    "cost_source": source,
                    "cost_updated_at": now,
                }},
            )
            matched += 1
        else:
            unmatched.append(entry.get("name") or f"cfid={entry.get('caffesta_product_id')}")

    return {"matched": matched, "total": len(entries), "unmatched": unmatched[:20]}


# ============ Endpoints ============
@router.post("/restaurants/{restaurant_id}/costs/upload")
async def upload_costs(
    restaurant_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    _: dict = Depends(ensure_can_write_system),
):
    await check_restaurant_access(current_user, restaurant_id)
    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            entries = _parse_xlsx(content)
        elif filename.endswith(".csv"):
            entries = _parse_csv(content)
        else:
            # Try autodetect — xlsx is zip, starts with PK
            if content[:2] == b"PK":
                entries = _parse_xlsx(content)
            else:
                entries = _parse_csv(content)
    except Exception as e:
        raise HTTPException(400, f"Не удалось разобрать файл: {e}")

    if not entries:
        raise HTTPException(400, "Файл пустой или не содержит валидных данных. Нужны колонки 'Название' и 'Себестоимость'.")

    result = await _match_and_update(restaurant_id, entries, source="file")
    await _send_margin_alerts(restaurant_id)
    return result


@router.post("/restaurants/{restaurant_id}/costs/import-caffesta")
async def import_caffesta_costs(restaurant_id: str, current_user: dict = Depends(get_current_user), _: dict = Depends(ensure_can_write_system)):
    await check_restaurant_access(current_user, restaurant_id)
    config = await get_caffesta_config(restaurant_id)
    if not config or not config.get("enabled") or not config.get("api_key"):
        raise HTTPException(400, "Caffesta не настроена для этого ресторана")

    balances = await caffesta_get_balances(restaurant_id)
    if not balances.get("ok"):
        raise HTTPException(502, f"Ошибка Caffesta API: {balances.get('message')}")

    entries = [
        {"caffesta_product_id": b["product_id"], "cost": b["self_cost"]}
        for b in balances.get("data", []) if b.get("self_cost", 0) > 0
    ]

    if not entries:
        return {"matched": 0, "total": 0, "unmatched": [], "message": "В Caffesta нет товаров с себестоимостью (self_cost=0 у всех)"}

    result = await _match_and_update(restaurant_id, entries, source="caffesta")
    await _send_margin_alerts(restaurant_id)
    return result


@router.get("/restaurants/{restaurant_id}/costs/analysis")
async def costs_analysis(restaurant_id: str, current_user: dict = Depends(get_current_user)):
    await check_restaurant_access(current_user, restaurant_id)
    settings = await get_or_create_settings(restaurant_id)
    items = await db.menu_items.find(
        {"restaurant_id": restaurant_id, "is_banner": {"$ne": True}}, {"_id": 0}
    ).to_list(10000)
    categories = await db.categories.find({"restaurant_id": restaurant_id}, {"_id": 0}).to_list(500)
    cat_map = {c["id"]: c for c in categories}

    result = []
    critical = 0
    warning = 0
    no_cost = 0
    total_margin_sum = 0.0
    total_margin_count = 0

    for item in items:
        category = cat_map.get(item.get("category_id"))
        threshold = _get_effective_threshold(item, category, settings)
        margin, special = _compute_margin(item.get("price"), item.get("cost_price"))

        if special == "no-cost":
            status = "no-cost"
            no_cost += 1
        elif margin is None:
            status = "no-price"
        elif margin < threshold - 5:
            status = "critical"
            critical += 1
        elif margin < threshold:
            status = "warning"
            warning += 1
        else:
            status = "ok"

        if margin is not None:
            total_margin_sum += margin
            total_margin_count += 1

        result.append({
            "id": item["id"],
            "name": item["name"],
            "category_id": item.get("category_id"),
            "category_name": category.get("name") if category else None,
            "price": item.get("price"),
            "cost_price": item.get("cost_price"),
            "cost_source": item.get("cost_source"),
            "cost_updated_at": item.get("cost_updated_at"),
            "margin_pct": margin,
            "threshold": threshold,
            "threshold_source": (
                "item" if item.get("margin_threshold") is not None else
                "category" if category and category.get("margin_threshold") is not None else
                "default"
            ),
            "status": status,
        })

    return {
        "items": result,
        "summary": {
            "total": len(items),
            "with_cost": len(items) - no_cost,
            "without_cost": no_cost,
            "critical": critical,
            "warning": warning,
            "ok": len(items) - no_cost - critical - warning,
            "avg_margin": round(total_margin_sum / total_margin_count, 1) if total_margin_count else None,
        },
    }


class CostUpdate(BaseModel):
    cost_price: Optional[float] = None
    margin_threshold: Optional[int] = None


@router.put("/restaurants/{restaurant_id}/menu-items/{item_id}/cost")
async def update_item_cost(
    restaurant_id: str, item_id: str, payload: CostUpdate,
    current_user: dict = Depends(get_current_user),
    _: dict = Depends(ensure_can_write_system),
):
    await check_restaurant_access(current_user, restaurant_id)
    update = {}
    if payload.cost_price is not None:
        update["cost_price"] = payload.cost_price
        update["cost_source"] = "manual"
        update["cost_updated_at"] = datetime.now(timezone.utc).isoformat()
    if payload.margin_threshold is not None:
        update["margin_threshold"] = payload.margin_threshold
    if not update:
        raise HTTPException(400, "Нечего обновлять")
    await db.menu_items.update_one({"id": item_id, "restaurant_id": restaurant_id}, {"$set": update})
    # Trigger alert check after manual edit (antispam handles duplicates)
    try:
        await _send_margin_alerts(restaurant_id)
    except Exception:
        pass
    return {"updated": True}


class CategoryThreshold(BaseModel):
    margin_threshold: Optional[int] = None


@router.put("/restaurants/{restaurant_id}/categories/{category_id}/threshold")
async def update_category_threshold(
    restaurant_id: str, category_id: str, payload: CategoryThreshold,
    current_user: dict = Depends(get_current_user),
    _: dict = Depends(ensure_can_write_system),
):
    await check_restaurant_access(current_user, restaurant_id)
    await db.categories.update_one(
        {"id": category_id, "restaurant_id": restaurant_id},
        {"$set": {"margin_threshold": payload.margin_threshold}},
    )
    return {"updated": True}


# ============ Telegram alerts ============
def _margin_signature(price, cost) -> str:
    """Signature used for antispam: if price+cost haven't changed, don't re-alert."""
    return f"{round(float(price or 0), 2)}:{round(float(cost or 0), 2)}"


async def _send_margin_alerts(restaurant_id: str, force: bool = False) -> dict:
    """Собирает блюда с критичной маржой и шлёт в Telegram если включены алерты.
    Антиспам: не шлёт повторно по блюду, если (price, cost_price) не изменились с прошлого алерта.
    force=True — игнорировать антиспам.
    """
    settings = await get_or_create_settings(restaurant_id)
    if not settings.get("margin_alerts_enabled"):
        return {"sent": False, "reason": "disabled"}
    token = (settings.get("margin_alerts_bot_token") or "").strip()
    chat_id = (settings.get("margin_alerts_chat_id") or "").strip()
    if not token or not chat_id:
        return {"sent": False, "reason": "no-credentials"}

    restaurant = await db.restaurants.find_one({"id": restaurant_id}, {"_id": 0, "name": 1})
    rest_name = restaurant.get("name", "Ресторан") if restaurant else "Ресторан"

    items = await db.menu_items.find(
        {"restaurant_id": restaurant_id, "is_banner": {"$ne": True}, "cost_price": {"$gt": 0}},
        {"_id": 0},
    ).to_list(10000)
    categories = await db.categories.find({"restaurant_id": restaurant_id}, {"_id": 0}).to_list(500)
    cat_map = {c["id"]: c for c in categories}

    critical = []
    skipped_antispam = 0
    for item in items:
        category = cat_map.get(item.get("category_id"))
        threshold = _get_effective_threshold(item, category, settings)
        margin, _ = _compute_margin(item.get("price"), item.get("cost_price"))
        if margin is None or margin >= threshold - 5:
            continue
        # Antispam: skip if signature unchanged
        sig = _margin_signature(item.get("price"), item.get("cost_price"))
        if not force and item.get("last_margin_alert_sig") == sig:
            skipped_antispam += 1
            continue
        critical.append({
            "id": item["id"],
            "name": item["name"],
            "price": item["price"],
            "cost": item["cost_price"],
            "margin": margin,
            "threshold": threshold,
            "sig": sig,
        })

    if not critical:
        return {"sent": False, "reason": "no-changes", "skipped": skipped_antispam}

    lines = [f"⚠️ <b>{rest_name}</b> — критическое падение маржи у {len(critical)} позиций:\n"]
    for it in critical[:20]:
        lines.append(
            f"• <b>{it['name']}</b> — цена {it['price']}, себест. {it['cost']}, "
            f"маржа <b>{it['margin']}%</b> (порог {it['threshold']}%)"
        )
    if len(critical) > 20:
        lines.append(f"\n...и ещё {len(critical) - 20}")
    text = "\n".join(lines)

    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.post(
                f"https://api.telegram.org/bot{token}/sendMessage",
                data={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
            )
            if resp.status_code >= 400:
                return {"sent": False, "reason": f"telegram-{resp.status_code}"}
    except Exception as e:
        return {"sent": False, "reason": f"exception:{e}"}

    # Update signatures for alerted items (antispam)
    now = datetime.now(timezone.utc).isoformat()
    for it in critical:
        await db.menu_items.update_one(
            {"id": it["id"]},
            {"$set": {"last_margin_alert_sig": it["sig"], "last_margin_alert_at": now}},
        )
    return {"sent": True, "count": len(critical), "skipped_antispam": skipped_antispam}


async def run_margin_check_job():
    """Daily scheduler job — check all restaurants with alerts enabled."""
    import logging
    cursor = db.restaurants.find({}, {"_id": 0, "id": 1, "name": 1})
    async for rest in cursor:
        try:
            settings = await db.settings.find_one({"restaurant_id": rest["id"]}, {"_id": 0})
            if not settings or not settings.get("margin_alerts_enabled"):
                continue
            result = await _send_margin_alerts(rest["id"])
            logging.info(f"Daily margin check {rest['name']}: {result}")
        except Exception as e:
            logging.exception(f"Margin check failed for {rest.get('name')}: {e}")


@router.post("/restaurants/{restaurant_id}/costs/check-alerts")
async def trigger_alerts(
    restaurant_id: str, force: bool = False,
    current_user: dict = Depends(get_current_user),
):
    await check_restaurant_access(current_user, restaurant_id)
    result = await _send_margin_alerts(restaurant_id, force=force)
    return result


@router.post("/restaurants/{restaurant_id}/costs/reset-alerts")
async def reset_alerts(restaurant_id: str, current_user: dict = Depends(get_current_user)):
    """Сбросить антиспам-историю — следующая проверка разошлёт алерты по всем критичным блюдам."""
    await check_restaurant_access(current_user, restaurant_id)
    result = await db.menu_items.update_many(
        {"restaurant_id": restaurant_id},
        {"$unset": {"last_margin_alert_sig": "", "last_margin_alert_at": ""}},
    )
    return {"reset": result.modified_count}


# ============ FACTUAL MARGIN FROM RECEIPTS ============

@router.get("/restaurants/{restaurant_id}/costs/factual-margin")
async def factual_margin(
    restaurant_id: str,
    days: int = 30,
    current_user: dict = Depends(get_current_user),
):
    """Compute FACTUAL margin by aggregating Caffesta receipts (uses self_cost_sum from each dish row).
    Returns per-product stats sorted by margin_pct ascending (lowest first)."""
    await check_restaurant_access(current_user, restaurant_id)
    if days < 1 or days > 120:
        raise HTTPException(400, "Период должен быть от 1 до 120 дней")

    from services.caffesta import caffesta_get_all_receipts, caffesta_get_products

    end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")

    res = await caffesta_get_all_receipts(restaurant_id, start_date, end_date)
    if not res.get("ok"):
        raise HTTPException(status_code=400, detail=res.get("message", "Ошибка Caffesta"))
    receipts = res.get("data", [])

    # product_id -> title
    product_map = {}
    prods = await caffesta_get_products(restaurant_id)
    if prods.get("ok"):
        for p in prods.get("data", []):
            try:
                product_map[int(p["product_id"])] = {"title": p.get("title") or "", "price": p.get("price")}
            except (ValueError, TypeError, KeyError):
                continue

    # Aggregate per product_id
    agg = {}  # pid -> {qty, revenue, cost}
    for r in receipts:
        if r.get("income") != 1:
            continue
        for od in (r.get("order_dishes") or []):
            dish = od.get("dish") or {}
            try:
                pid = int(dish.get("id") or 0)
            except (ValueError, TypeError):
                pid = 0
            if not pid:
                continue
            try:
                qty = float(od.get("count") or 1)
                rev = float(od.get("total_sum") or 0)
                cost = float(od.get("self_cost_sum") or 0)
            except (ValueError, TypeError):
                continue
            agg.setdefault(pid, {"pid": pid, "qty": 0, "revenue": 0.0, "cost": 0.0})
            agg[pid]["qty"] += qty
            agg[pid]["revenue"] += rev
            agg[pid]["cost"] += cost

    # Build items with margin
    items = []
    total_rev = 0.0
    total_cost = 0.0
    for pid, a in agg.items():
        pinfo = product_map.get(pid) or {}
        rev = a["revenue"]
        cost = a["cost"]
        margin_abs = rev - cost
        margin_pct = round((margin_abs / rev * 100), 1) if rev > 0 else None
        items.append({
            "product_id": pid,
            "title": pinfo.get("title") or f"ID #{pid}",
            "catalog_price": pinfo.get("price"),
            "qty": round(a["qty"], 2),
            "revenue": round(rev, 2),
            "cost": round(cost, 2),
            "margin_abs": round(margin_abs, 2),
            "margin_pct": margin_pct,
            "avg_price": round(rev / a["qty"], 2) if a["qty"] else None,
            "avg_cost": round(cost / a["qty"], 2) if a["qty"] else None,
        })
        total_rev += rev
        total_cost += cost

    # Sort: lowest margin% first (worst at top); items with None margin at end
    items.sort(key=lambda x: (x["margin_pct"] is None, x["margin_pct"] if x["margin_pct"] is not None else 999))

    total_margin_abs = total_rev - total_cost
    total_margin_pct = round((total_margin_abs / total_rev * 100), 1) if total_rev > 0 else 0

    return {
        "period": {"start": start_date, "end": end_date, "days": days},
        "summary": {
            "products_count": len(items),
            "total_revenue": round(total_rev, 2),
            "total_cost": round(total_cost, 2),
            "total_margin_abs": round(total_margin_abs, 2),
            "total_margin_pct": total_margin_pct,
        },
        "items": items,
    }
